from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models.base import db
from models.solicitud import Solicitud          
from models.artesano import Artesano            
from models.rubro import Rubro                  
from models.estado_solicitud import EstadoSolicitud 
from models.usuario import Usuario              
from models.administrador import Administrador
from models.solicitud_foto import SolicitudFoto
from models.notificacion import Notificacion
from models.limite_rubro import LimiteRubro
from models.mapa import Mapa  
from models.parcela import Parcela  
from models.solicitud_parcela import SolicitudParcela 
from models.color import Color
from sqlalchemy import or_, func, and_
from datetime import datetime, timedelta
from sqlalchemy.orm import joinedload 
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Blueprint con prefix
admin_bp = Blueprint('admin', __name__, url_prefix='/api/v1') 

def get_usuario_actual():
    """Obtiene el usuario actual desde el token JWT"""
    user_identity = get_jwt_identity()
    usuario_id = int(user_identity.split('_')[1])
    return Usuario.query.get(usuario_id)

def get_administrador_actual():
    """Obtiene el administrador actual basado en el usuario del token"""
    usuario = get_usuario_actual()
    if not usuario:
        return None
    return Administrador.query.filter_by(usuario_id=usuario.usuario_id).first()

class AdminController:

    @staticmethod
    def _check_admin_permissions(usuario):
        """Verifica si el usuario es Administrador (2) u Organizador (3)."""
        if not usuario:
            return {'msg': 'Usuario no encontrado'}, 404
            
        if usuario.rol_id not in [2, 3]:
            return {'msg': 'Acceso denegado. Se requiere rol de Administrador u Organizador.'}, 403
            
        administrador = Administrador.query.filter_by(usuario_id=usuario.usuario_id).first()
        if not administrador:
            return {'msg': 'Perfil de administrador no encontrado'}, 404
            
        return administrador

    @staticmethod
    def crear_notificacion_artesano(artesano_id, mensaje, estado_notificacion_id=1):
        """
        Crear una nueva notificación para un artesano
        """
        try:
            notificacion = Notificacion(
                artesano_id=artesano_id,
                mensaje=mensaje,
                estado_notificacion_id=estado_notificacion_id
            )
            db.session.add(notificacion)
            return notificacion
        except Exception as e:
            print(f"Error al crear notificación: {str(e)}")
            raise e

    @staticmethod
    def get_limite_activo_rubro(rubro_id):
        """
        Obtiene el límite activo actual para un rubro
        """
        try:
            limite_activo = LimiteRubro.query.filter_by(
                rubro_id=rubro_id,
                es_activo=True
            ).first()
            
            if limite_activo:
                return limite_activo.max_puestos, limite_activo.limite_id
            return None, None
            
        except Exception as e:
            print(f"Error al obtener límite activo para rubro {rubro_id}: {str(e)}")
            return None, None

    @staticmethod
    def verificar_limite_rubro(rubro_id):
        """
        Verifica si se ha alcanzado el límite máximo de puestos para un rubro - ACTUALIZADO
        """
        try:
            # Obtener el límite activo para el rubro
            limite_maximo, limite_id = AdminController.get_limite_activo_rubro(rubro_id)
            
            if not limite_maximo:
                return False, 0, 0  # Sin límite configurado

            # Contar solicitudes aprobadas para este rubro
            estado_aprobada = EstadoSolicitud.query.filter_by(nombre='Aprobada').first()
            count_aprobadas = db.session.query(func.count(Solicitud.solicitud_id)).filter(
                Solicitud.rubro_id == rubro_id,
                Solicitud.estado_solicitud_id == estado_aprobada.estado_solicitud_id
            ).scalar() or 0

            limite_alcanzado = count_aprobadas == limite_maximo
            
            return limite_alcanzado, count_aprobadas, limite_maximo

        except Exception as e:
            print(f"Error al verificar límite de rubro: {str(e)}")
            return False, 0, 0

    @staticmethod
    def get_solicitudes_dashboard(filtro_estado=None, busqueda_termino=None):
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return permisos

        try:
            query = Solicitud.query.options(
                joinedload(Solicitud.estado_rel),
                joinedload(Solicitud.rubro_rel),
                joinedload(Solicitud.artesano_rel)
            )
            query = query.join(Solicitud.estado_rel).join(Solicitud.rubro_rel).join(Solicitud.artesano_rel)

            if filtro_estado and filtro_estado != 'all':
                query = query.filter(EstadoSolicitud.nombre == filtro_estado)

            if busqueda_termino:
                term = f"%{busqueda_termino.lower()}%"
                query = query.filter(or_(
                    db.cast(Solicitud.solicitud_id, db.String).ilike(term),
                    db.func.lower(Artesano.nombre).like(term), 
                    db.func.lower(Rubro.tipo).like(term)      
                ))

            solicitudes = query.order_by(Solicitud.fecha_solicitud.desc()).all()
            
            data = []
            for s in solicitudes: 
                artesano = s.artesano_rel
                
                foto_principal = s.fotos_rel[0] if s.fotos_rel else None
                foto_url = foto_principal.get_image_url() if foto_principal else None
                
                # Verificar límite de rubro para esta solicitud
                limite_alcanzado, count_actual, limite_maximo = AdminController.verificar_limite_rubro(s.rubro_id)
                
                original_data = {
                    'solicitud_id': s.solicitud_id, 
                    'nombre': artesano.nombre if artesano else 'Sin nombre',
                    'rubro': s.rubro_rel.tipo if s.rubro_rel else 'No especificado',
                    'rubro_id': s.rubro_id,
                    'alto': float(s.dimensiones_largo) if s.dimensiones_largo else None, 
                    'ancho': float(s.dimensiones_ancho) if s.dimensiones_ancho else None, 
                    'descripcion_puesto': s.descripcion or 'Sin descripción proporcionada',
                    'estado_solicitud': s.estado_rel.nombre if s.estado_rel else 'Pendiente',
                    'fecha_solicitud': s.fecha_solicitud.isoformat() if s.fecha_solicitud else None,
                    'notas_admin': s.comentarios_admin or "",
                    'foto_puesto': foto_url,
                    'fotos': [foto.get_image_url() for foto in s.fotos_rel] if s.fotos_rel else [],
                    'limite_rubro_alcanzado': limite_alcanzado,
                    'count_actual_rubro': count_actual,
                    'limite_maximo_rubro': limite_maximo
                }
                
                if artesano:
                    if hasattr(artesano, 'telefono'):
                        original_data['telefono'] = artesano.telefono
                    else:
                        original_data['telefono'] = 'No especificado'
                    
                    if hasattr(artesano, 'dni'):
                        original_data['dni'] = artesano.dni
                    else:
                        original_data['dni'] = 'No especificado'
                    
                    original_data['email'] = 'No disponible'
                else:
                    original_data['telefono'] = 'No especificado'
                    original_data['dni'] = 'No especificado'
                    original_data['email'] = 'No disponible'
                
                data.append({
                    'id': s.solicitud_id,
                    'nombre': artesano.nombre if artesano else 'Sin nombre',
                    'rubro': s.rubro_rel.tipo if s.rubro_rel else 'No especificado',
                    'rubro_id': s.rubro_id,
                    'dimensiones': f"{float(s.dimensiones_largo)}x{float(s.dimensiones_ancho)}", 
                    'estado': s.estado_rel.nombre if s.estado_rel else 'Pendiente',
                    'fechaSolicitud': s.fecha_solicitud.isoformat() if s.fecha_solicitud else None,
                    'artesano_id': artesano.artesano_id if artesano else None,
                    'originalData': original_data,
                    'limite_rubro_alcanzado': limite_alcanzado
                })
            
            return data, 200

        except Exception as e:
            print(f"Error al obtener solicitudes: {str(e)}")
            return {'msg': 'Error interno al obtener las solicitudes.', 'detalle': str(e)}, 500

    @staticmethod
    def actualizar_estado_solicitud(solicitud_id, data):
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return permisos
            
        estado_nombre_nuevo = data.get('estado_solicitud')
        comentarios_admin = data.get('notas_admin')
        administrador = permisos
        
        print(f"DEBUG: Actualizando estado de solicitud {solicitud_id}")
        print(f"DEBUG - Estado nuevo: {estado_nombre_nuevo}")
        print(f"DEBUG - Comentarios: {comentarios_admin}")
        
        # Validar que el estado esté presente
        if not estado_nombre_nuevo:
            return {'msg': 'El campo "estado_solicitud" es requerido.'}, 400

        nuevo_estado = EstadoSolicitud.query.filter_by(nombre=estado_nombre_nuevo).first()
        if not nuevo_estado:
            print(f"DEBUG: Estado '{estado_nombre_nuevo}' no encontrado en la base de datos")
            return {'msg': f'El estado "{estado_nombre_nuevo}" no es válido.'}, 400

        try:
            solicitud = Solicitud.query.get(solicitud_id)
            if not solicitud: 
                return {'msg': f'Solicitud ID {solicitud_id} no encontrada.'}, 404

            estado_anterior = solicitud.estado_rel.nombre if solicitud.estado_rel else 'Sin estado'
            
            print(f"DEBUG: Estado anterior: {estado_anterior}")
            
            # Verificar límite si se está aprobando
            if estado_nombre_nuevo == 'Aprobada':
                print("DEBUG: Verificando límite de rubro para aprobación")
                limite_alcanzado, count_actual, limite_maximo = AdminController.verificar_limite_rubro(solicitud.rubro_id)
                if limite_alcanzado:
                    return {
                        'msg': f'No se puede aprobar la solicitud. Límite de puestos alcanzado para este rubro ({count_actual}/{limite_maximo}).',
                        'limite_alcanzado': True,
                        'count_actual': count_actual,
                        'limite_maximo': limite_maximo
                    }, 400

            # Actualizar el estado
            solicitud.estado_solicitud_id = nuevo_estado.estado_solicitud_id
            solicitud.comentarios_admin = comentarios_admin
            solicitud.administrador_id = administrador.administrador_id 
            solicitud.fecha_gestion = datetime.utcnow()

            # Crear notificación para el artesano
            if solicitud.artesano_id:
                mensaje_notificacion = f"El estado de tu solicitud cambió de '{estado_anterior}' a '{estado_nombre_nuevo}'."
                
                if comentarios_admin:
                    mensaje_notificacion += f" Comentarios del administrador: {comentarios_admin}"
                
                AdminController.crear_notificacion_artesano(
                    solicitud.artesano_id, 
                    mensaje_notificacion
                )

            db.session.commit()
            print(f"DEBUG: Solicitud {solicitud_id} actualizada exitosamente a {estado_nombre_nuevo}")
            return {'msg': f'Estado de la solicitud {solicitud_id} actualizado a {estado_nombre_nuevo}'}, 200

        except Exception as e:
            db.session.rollback()
            print(f"DEBUG: Error al actualizar solicitud: {str(e)}")
            return {'msg': 'Error interno al actualizar la solicitud.', 'error': str(e)}, 500

    @staticmethod
    def modificar_informacion_puesto(solicitud_id, data):
        """
        RF13: Modificar información del puesto y notificar al artesano
        """
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return permisos

        try:
            solicitud = Solicitud.query.get(solicitud_id)
            if not solicitud:
                return {'msg': f'Solicitud ID {solicitud_id} no encontrada.'}, 404

            cambios = []
            
            # Verificar y actualizar rubro
            if 'rubro_id' in data and data['rubro_id'] != solicitud.rubro_id:
                nuevo_rubro = Rubro.query.get(data['rubro_id'])
                if not nuevo_rubro:
                    return {'msg': 'Rubro no válido'}, 400
                
                rubro_anterior = solicitud.rubro_rel.tipo if solicitud.rubro_rel else 'No especificado'
                solicitud.rubro_id = data['rubro_id']
                cambios.append(f"rubro de '{rubro_anterior}' a '{nuevo_rubro.tipo}'")

            # Verificar y actualizar dimensiones
            if 'dimensiones_largo' in data and data['dimensiones_largo'] != solicitud.dimensiones_largo:
                largo_anterior = solicitud.dimensiones_largo
                solicitud.dimensiones_largo = data['dimensiones_largo']
                cambios.append(f"largo de {largo_anterior}m a {data['dimensiones_largo']}m")

            if 'dimensiones_ancho' in data and data['dimensiones_ancho'] != solicitud.dimensiones_ancho:
                ancho_anterior = solicitud.dimensiones_ancho
                solicitud.dimensiones_ancho = data['dimensiones_ancho']
                cambios.append(f"ancho de {ancho_anterior}m a {data['dimensiones_ancho']}m")

            # Verificar y actualizar descripción
            if 'descripcion' in data and data['descripcion'] != solicitud.descripcion:
                solicitud.descripcion = data['descripcion']
                cambios.append("descripción del puesto")

            # Verificar y actualizar comentarios admin
            if 'comentarios_admin' in data:
                solicitud.comentarios_admin = data['comentarios_admin']

            solicitud.administrador_id = permisos.administrador_id
            solicitud.fecha_gestion = datetime.utcnow()

            # Crear notificación para el artesano si hay cambios
            if cambios and solicitud.artesano_id:
                mensaje_cambios = ", ".join(cambios)
                mensaje_notificacion = f"El administrador ha modificado la siguiente información de tu solicitud: {mensaje_cambios}."
                
                if data.get('comentarios_admin'):
                    mensaje_notificacion += f" Motivo: {data['comentarios_admin']}"
                
                AdminController.crear_notificacion_artesano(
                    solicitud.artesano_id,
                    mensaje_notificacion
                )

            db.session.commit()

            return {
                'msg': f'Información del puesto actualizada correctamente.',
                'cambios': cambios
            }, 200

        except Exception as e:
            db.session.rollback()
            return {'msg': 'Error interno al modificar la información del puesto.', 'error': str(e)}, 500

    @staticmethod
    def get_configuraciones_rubros():
        """
        RF17: Obtener configuraciones de precios y límites por rubro - ACTUALIZADO
        """
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return permisos

        try:
            rubros = Rubro.query.all()
            configuraciones = []

            for rubro in rubros:
                # Obtener límite activo para el rubro
                limite_activo, limite_id = AdminController.get_limite_activo_rubro(rubro.rubro_id)
                
                # Contar puestos aprobados
                estado_aprobada = EstadoSolicitud.query.filter_by(nombre='Aprobada').first()
                count_aprobadas = db.session.query(func.count(Solicitud.solicitud_id)).filter(
                    Solicitud.rubro_id == rubro.rubro_id,
                    Solicitud.estado_solicitud_id == estado_aprobada.estado_solicitud_id
                ).scalar() or 0

                # Verificar disponibilidad
                disponible = True
                if limite_activo and count_aprobadas >= limite_activo:
                    disponible = False

                configuraciones.append({
                    'rubro_id': rubro.rubro_id,
                    'rubro_nombre': rubro.tipo,
                    'precio_base': float(rubro.precio_parcela) if rubro.precio_parcela else 0.0,
                    'limite_puestos': limite_activo,  # Usa el límite de la tabla LimiteRubro
                    'limite_id': limite_id,  # ID del límite para actualizaciones
                    'puestos_aprobados': count_aprobadas,
                    'disponible': disponible
                })

            return configuraciones, 200

        except Exception as e:
            print(f"Error al obtener configuraciones de rubros: {str(e)}")
            return {'msg': 'Error interno al obtener las configuraciones.', 'detalle': str(e)}, 500

    @staticmethod
    def actualizar_configuracion_rubro(rubro_id, data):
        """
        RF17: Actualizar configuración de precio y límite para un rubro - ACTUALIZADO
        """
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return permisos

        try:
            rubro = Rubro.query.get(rubro_id)
            if not rubro:
                return {'msg': 'Rubro no encontrado'}, 404

            # Actualizar precio en el rubro
            if 'precio_base' in data:
                rubro.precio_parcela = data['precio_base']

            # Actualizar o crear límite en LimiteRubro
            if 'limite_puestos' in data:
                limite_activo, limite_id = AdminController.get_limite_activo_rubro(rubro_id)
                
                nuevo_limite = data['limite_puestos']
                
                if limite_id:  # Actualizar límite existente
                    limite = LimiteRubro.query.get(limite_id)
                    if nuevo_limite:
                        limite.max_puestos = nuevo_limite
                        limite.fecha_vigencia = datetime.utcnow().date()
                    else:
                        # Si se establece None, desactivar el límite
                        limite.es_activo = False
                else:  # Crear nuevo límite
                    if nuevo_limite:
                        nuevo_limite_rubro = LimiteRubro(
                            rubro_id=rubro_id,
                            max_puestos=nuevo_limite,
                            fecha_vigencia=datetime.utcnow().date(),
                            es_activo=True
                        )
                        db.session.add(nuevo_limite_rubro)

            db.session.commit()

            return {
                'msg': f'Configuración del rubro {rubro.tipo} actualizada correctamente.',
                'configuracion': {
                    'rubro_id': rubro.rubro_id,
                    'rubro_nombre': rubro.tipo,
                    'precio_base': float(rubro.precio_parcela) if rubro.precio_parcela else 0.0,
                    'limite_puestos': data.get('limite_puestos')
                }
            }, 200

        except Exception as e:
            db.session.rollback()
            return {'msg': 'Error interno al actualizar la configuración.', 'error': str(e)}, 500

    @staticmethod
    def get_diversidad_rubros():
        """
        RF14: Obtener diversidad por categorías con límites - ACTUALIZADO
        """
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return permisos

        try:
            rubros_activos = Rubro.query.all()
            
            resultado = []
            for rubro in rubros_activos:
                # Contar por estado para ESTE rubro específico
                total_solicitudes = db.session.query(func.count(Solicitud.solicitud_id)).filter(
                    Solicitud.rubro_id == rubro.rubro_id
                ).scalar() or 0
                
                estado_aprobada = EstadoSolicitud.query.filter_by(nombre='Aprobada').first()
                aprobadas = db.session.query(func.count(Solicitud.solicitud_id)).filter(
                    Solicitud.rubro_id == rubro.rubro_id,
                    Solicitud.estado_solicitud_id == estado_aprobada.estado_solicitud_id
                ).scalar() or 0
                
                estados_pendientes = EstadoSolicitud.query.filter(
                    EstadoSolicitud.nombre.in_(['Pendiente', 'Pendiente por Modificación'])
                ).all()
                estados_pendientes_ids = [e.estado_solicitud_id for e in estados_pendientes]
                
                pendientes = db.session.query(func.count(Solicitud.solicitud_id)).filter(
                    Solicitud.rubro_id == rubro.rubro_id,
                    Solicitud.estado_solicitud_id.in_(estados_pendientes_ids)
                ).scalar() or 0
                
                # Obtener límite activo
                limite_activo, _ = AdminController.get_limite_activo_rubro(rubro.rubro_id)
                
                # Calcular disponibilidad
                limite_alcanzado = False
                disponibilidad = "Sin límite"
                
                if limite_activo:
                    limite_alcanzado = aprobadas >= limite_activo
                    disponibilidad = f"{aprobadas}/{limite_activo}"
                
                resultado.append({
                    'rubro_id': rubro.rubro_id,
                    'rubro_nombre': rubro.tipo,
                    'total_solicitudes': total_solicitudes,
                    'aprobadas': aprobadas,
                    'pendientes': pendientes,
                    'limite_puestos': limite_activo,
                    'precio_base': float(rubro.precio_parcela) if rubro.precio_parcela else 0.0,
                    'limite_alcanzado': limite_alcanzado,
                    'disponibilidad': disponibilidad
                })

            return resultado, 200

        except Exception as e:
            print(f"Error al obtener diversidad de rubros: {str(e)}")
            return {'msg': 'Error interno al obtener la diversidad de rubros.', 'detalle': str(e)}, 500

    @staticmethod
    def get_estadisticas_usuarios(fecha_inicio, fecha_fin, agrupacion='dia'):
        """
        Obtener estadísticas de usuarios por período - ACTUALIZADO CON VALIDACIONES Y CONSULTAS COMPATIBLES
        """
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return permisos

        try:
            # Validaciones de fecha
            hoy = datetime.utcnow().date()
            
            # Validar que fecha_inicio no sea mayor que fecha_fin
            if fecha_inicio > fecha_fin:
                return {'msg': 'La fecha de inicio no puede ser mayor que la fecha de fin'}, 400
            
            # Validar que no se seleccionen fechas futuras
            if fecha_inicio > hoy or fecha_fin > hoy:
                return {'msg': 'No se pueden seleccionar fechas futuras'}, 400
            
            # Validar que el rango no sea excesivamente grande (máximo 1 año)
            if (fecha_fin - fecha_inicio).days > 365:
                return {'msg': 'El rango de fechas no puede ser mayor a 1 año'}, 400

            # Consultas compatibles con ONLY_FULL_GROUP_BY
            if agrupacion == 'semana':
                # Para semana: usar la misma expresión en SELECT y GROUP BY
                query = db.session.query(
                    func.yearweek(Usuario.fecha_registro).label('periodo'),
                    func.count(Usuario.usuario_id).label('nuevos_usuarios')
                ).filter(
                    Usuario.fecha_registro.between(fecha_inicio, fecha_fin + timedelta(days=1))
                ).group_by(
                    func.yearweek(Usuario.fecha_registro)
                ).order_by('periodo')
                
                resultados = query.all()
                
                # Formatear resultados para semanas
                estadisticas = []
                for resultado in resultados:
                    # Convertir año-semana a formato legible
                    año_semana = str(resultado.periodo)
                    año = año_semana[:4]
                    semana = año_semana[4:]
                    fecha_str = f"Semana {semana}-{año}"
                    
                    estadisticas.append({
                        'fecha': fecha_str,
                        'nuevos_usuarios': resultado.nuevos_usuarios
                    })

            elif agrupacion == 'mes':
                # Para mes: usar las mismas expresiones en SELECT y GROUP BY
                query = db.session.query(
                    func.year(Usuario.fecha_registro).label('año'),
                    func.month(Usuario.fecha_registro).label('mes'),
                    func.count(Usuario.usuario_id).label('nuevos_usuarios')
                ).filter(
                    Usuario.fecha_registro.between(fecha_inicio, fecha_fin + timedelta(days=1))
                ).group_by(
                    func.year(Usuario.fecha_registro),
                    func.month(Usuario.fecha_registro)
                ).order_by('año', 'mes')
                
                resultados = query.all()
                
                # Formatear resultados para meses
                estadisticas = []
                for resultado in resultados:
                    fecha_str = f"{resultado.año}-{resultado.mes:02d}"
                    
                    estadisticas.append({
                        'fecha': fecha_str,
                        'nuevos_usuarios': resultado.nuevos_usuarios
                    })

            else:  # día por defecto
                # Para día: usar la misma expresión en SELECT y GROUP BY
                query = db.session.query(
                    func.date(Usuario.fecha_registro).label('fecha'),
                    func.count(Usuario.usuario_id).label('nuevos_usuarios')
                ).filter(
                    Usuario.fecha_registro.between(fecha_inicio, fecha_fin + timedelta(days=1))
                ).group_by(
                    func.date(Usuario.fecha_registro)
                ).order_by('fecha')
                
                resultados = query.all()
                
                # Formatear resultados para días
                estadisticas = []
                for resultado in resultados:
                    estadisticas.append({
                        'fecha': resultado.fecha.strftime('%Y-%m-%d'),
                        'nuevos_usuarios': resultado.nuevos_usuarios
                    })

            # Calcular total general
            total_general = sum(item['nuevos_usuarios'] for item in estadisticas)

            return {
                'estadisticas': estadisticas,
                'total_general': total_general,
                'rango_fechas': {
                    'inicio': fecha_inicio.strftime('%Y-%m-%d'),
                    'fin': fecha_fin.strftime('%Y-%m-%d')
                },
                'agrupacion': agrupacion
            }, 200

        except Exception as e:
            print(f"Error al obtener estadísticas de usuarios: {str(e)}")
            return {'msg': 'Error interno al obtener las estadísticas.', 'detalle': str(e)}, 500

    @staticmethod
    def exportar_estadisticas_excel(fecha_inicio, fecha_fin, agrupacion='dia'):
        """
        Exportar estadísticas de usuarios a Excel - ACTUALIZADO CON CONSULTAS COMPATIBLES
        """
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return jsonify(permisos[0]), permisos[1]

        try:
            # Validaciones de fecha
            hoy = datetime.utcnow().date()
            
            if fecha_inicio > fecha_fin:
                return {'msg': 'La fecha de inicio no puede ser mayor que la fecha de fin'}, 400
            
            if fecha_inicio > hoy or fecha_fin > hoy:
                return {'msg': 'No se pueden seleccionar fechas futuras'}, 400

            # Obtener datos usando el método principal
            resultado, status = AdminController.get_estadisticas_usuarios(fecha_inicio, fecha_fin, agrupacion)
            if status != 200:
                return resultado, status

            estadisticas = resultado['estadisticas']
            total_general = resultado['total_general']

            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Estadísticas Usuarios"

            # Título y metadatos
            ws['A1'] = f"Reporte de Usuarios Registrados"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Período: {fecha_inicio} a {fecha_fin}"
            ws['A3'] = f"Agrupación: {agrupacion}"
            ws['A4'] = f"Total de usuarios: {total_general}"

            # Encabezados
            headers = ['Fecha/Período', 'Nuevos Usuarios']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Datos
            for row, stat in enumerate(estadisticas, 7):
                ws.cell(row=row, column=1, value=stat['fecha'])
                ws.cell(row=row, column=2, value=stat['nuevos_usuarios'])

            # Total al final
            total_row = len(estadisticas) + 7
            ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
            ws.cell(row=total_row, column=2, value=total_general).font = Font(bold=True)

            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15

            # Guardar en buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return buffer, 200

        except Exception as e:
            print(f"Error al exportar estadísticas a Excel: {str(e)}")
            return {'msg': 'Error interno al exportar las estadísticas.', 'detalle': str(e)}, 500

    @staticmethod
    def exportar_artesanos_pdf():
        """
        Exportar listado de artesanos aprobados a PDF
        """
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return jsonify(permisos[0]), permisos[1]

        try:
            # Obtener artesanos aprobados
            estado_aprobada = EstadoSolicitud.query.filter_by(nombre='Aprobada').first()
            
            artesanos_aprobados = db.session.query(
                Artesano.nombre,
                Artesano.dni,
                Artesano.telefono,
                Rubro.tipo.label('rubro'),
                Solicitud.descripcion
            ).join(Solicitud, Solicitud.artesano_id == Artesano.artesano_id
            ).join(Rubro, Rubro.rubro_id == Solicitud.rubro_id
            ).filter(Solicitud.estado_solicitud_id == estado_aprobada.estado_solicitud_id
            ).order_by(Artesano.nombre).all()

            # Crear PDF
            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter

            # Título
            p.setFont("Helvetica-Bold", 16)
            p.drawString(100, height - 100, "Listado de Artesanos Aprobados")
            p.setFont("Helvetica", 10)
            p.drawString(100, height - 120, f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

            # Encabezados de tabla
            y_position = height - 160
            headers = ['Nombre', 'DNI', 'Teléfono', 'Rubro', 'Descripción']
            p.setFont("Helvetica-Bold", 10)
            
            # Dibujar encabezados
            p.drawString(50, y_position, headers[0])
            p.drawString(150, y_position, headers[1])
            p.drawString(220, y_position, headers[2])
            p.drawString(300, y_position, headers[3])
            p.drawString(400, y_position, headers[4])

            # Línea separadora
            p.line(50, y_position - 5, width - 50, y_position - 5)

            # Datos
            p.setFont("Helvetica", 8)
            y_position -= 20

            for artesano in artesanos_aprobados:
                # Verificar si necesita nueva página
                if y_position < 100:
                    p.showPage()
                    p.setFont("Helvetica", 8)
                    y_position = height - 50

                # Dibujar datos (truncar si es muy largo)
                nombre = artesano.nombre[:20] + '...' if len(artesano.nombre) > 20 else artesano.nombre
                descripcion = artesano.descripcion[:30] + '...' if artesano.descripcion and len(artesano.descripcion) > 30 else artesano.descripcion

                p.drawString(50, y_position, nombre)
                p.drawString(150, y_position, artesano.dni or 'N/A')
                p.drawString(220, y_position, artesano.telefono or 'N/A')
                p.drawString(300, y_position, artesano.rubro)
                p.drawString(400, y_position, descripcion or 'N/A')

                y_position -= 15

            p.save()
            buffer.seek(0)

            return buffer, 200

        except Exception as e:
            print(f"Error al exportar artesanos a PDF: {str(e)}")
            return {'msg': 'Error interno al exportar el listado de artesanos.', 'detalle': str(e)}, 500

    @staticmethod
    def get_estadisticas_rubros():
        """
        Obtener estadísticas de rubros (solo aprobadas)
        """
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return permisos

        try:
            estadisticas = db.session.query(
                Rubro.tipo.label('rubro'),
                func.count(Solicitud.solicitud_id).label('total')
            ).join(Solicitud, Solicitud.rubro_id == Rubro.rubro_id
            ).join(EstadoSolicitud, EstadoSolicitud.estado_solicitud_id == Solicitud.estado_solicitud_id
            ).filter(EstadoSolicitud.nombre == 'Aprobada'
            ).group_by(Rubro.tipo).all()
            
            resultado = {est.rubro: est.total for est in estadisticas}
            return resultado, 200

        except Exception as e:
            print(f"Error al obtener estadísticas de rubros: {str(e)}")
            return {'msg': 'Error interno al obtener las estadísticas de rubros.', 'detalle': str(e)}, 500

    @staticmethod
    def get_estadisticas_rubros_todas():
        """
        Obtener estadísticas de rubros (todas las solicitudes)
        """
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return permisos

        try:
            estadisticas = db.session.query(
                Rubro.tipo.label('rubro'),
                func.count(Solicitud.solicitud_id).label('total')
            ).join(Solicitud, Solicitud.rubro_id == Rubro.rubro_id
            ).group_by(Rubro.tipo).all()
            
            resultado = {est.rubro: est.total for est in estadisticas}
            return resultado, 200

        except Exception as e:
            print(f"Error al obtener estadísticas de rubros (todas): {str(e)}")
            return {'msg': 'Error interno al obtener las estadísticas de rubros.', 'detalle': str(e)}, 500

# RUTAS NUEVAS PARA LOS REQUERIMIENTOS

@admin_bp.route('/admin/parcelas', methods=['GET'])
@jwt_required()
def obtener_parcelas_admin():
    """Obtener parcelas para admin - VERSIÓN SIMPLIFICADA Y ROBUSTA"""
    try:
        print(" INICIANDO obtener_parcelas_admin")
        
        # Obtener identity DIRECTAMENTE
        user_identity = get_jwt_identity()
        print(f" User identity: {user_identity}")
        
        # Extraer ID del usuario - MANERA DIRECTA
        if isinstance(user_identity, str) and user_identity.startswith('user_'):
            usuario_id = int(user_identity.split('_')[1])
        else:
            usuario_id = int(user_identity)
            
        print(f" Usuario ID extraído: {usuario_id}")
        
        # Verificar si el usuario existe y es admin
        usuario = Usuario.query.get(usuario_id)
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 404
            
        print(f"🎯 Usuario encontrado: {usuario.email}, Rol: {usuario.rol_id}")
        
        # Verificar si es administrador u organizador
        if usuario.rol_id not in [2, 3]:
            return jsonify({'error': 'Acceso denegado. Se requiere rol de administrador.'}), 403
        
        # Verificar perfil de administrador
        administrador = Administrador.query.filter_by(usuario_id=usuario_id).first()
        if not administrador:
            return jsonify({'error': 'Perfil de administrador no encontrado'}), 404
            
        print(f"🎯 Administrador autorizado: {administrador.nombre}")
        
        # OBTENER DATOS DEL MAPA - MANERA DIRECTA
        mapa = Mapa.query.first()
        if not mapa:
            return jsonify({'error': 'No se ha configurado el mapa'}), 404
            
        print(f"🎯 Mapa encontrado: {mapa.mapa_id} - {mapa.cant_total_filas}x{mapa.cant_total_columnas}")
        
        # Obtener parcelas
        parcelas = Parcela.query.filter_by(mapa_id=mapa.mapa_id).all()
        print(f"🎯 Parcelas encontradas: {len(parcelas)}")
        
        # Procesar parcelas
        parcelas_data = []
        for parcela in parcelas:
            parcela_data = {
                'parcela_id': parcela.parcela_id,
                'fila': parcela.fila,
                'columna': parcela.columna,
                'habilitada': parcela.habilitada,
                'rubro_id': parcela.rubro_id,
                'mapa_id': parcela.mapa_id,
                'tipo_parcela_id': parcela.tipo_parcela_id
            }
            
            # Obtener información del rubro DIRECTAMENTE
            rubro = Rubro.query.get(parcela.rubro_id)
            if rubro:
                color = Color.query.get(rubro.color_id)
                parcela_data['rubro_info'] = {
                    'tipo': rubro.tipo,
                    'color': color.codigo_hex if color else '#CCCCCC'
                }
            
            # Verificar si está ocupada - CONSULTA DIRECTA
            solicitud_ocupada = db.session.query(SolicitudParcela).join(
                Solicitud, SolicitudParcela.solicitud_id == Solicitud.solicitud_id
            ).join(
                EstadoSolicitud, Solicitud.estado_solicitud_id == EstadoSolicitud.estado_solicitud_id
            ).filter(
                SolicitudParcela.parcela_id == parcela.parcela_id,
                EstadoSolicitud.nombre == 'Aprobada'
            ).first()
            
            parcela_data['ocupada'] = solicitud_ocupada is not None
            
            # Si está ocupada, obtener info del artesano
            if solicitud_ocupada:
                solicitud = Solicitud.query.get(solicitud_ocupada.solicitud_id)
                if solicitud and solicitud.artesano_id:
                    artesano = Artesano.query.get(solicitud.artesano_id)
                    if artesano:
                        parcela_data['artesano_info'] = {
                            'artesano_id': artesano.artesano_id,
                            'nombre': artesano.nombre,
                            'dni': artesano.dni,
                            'telefono': artesano.telefono
                        }
            
            parcelas_data.append(parcela_data)
        
        print("🎯 FINALIZADO EXITOSAMENTE")
        
        return jsonify({
            'parcelas': parcelas_data,
            'mapa': {
                'mapa_id': mapa.mapa_id,
                'cant_total_filas': mapa.cant_total_filas,
                'cant_total_columnas': mapa.cant_total_columnas
            },
            'total': len(parcelas_data)
        }), 200
        
    except Exception as e:
        import traceback
        print(f"❌ ERROR CRÍTICO en obtener_parcelas_admin: {str(e)}")
        print("❌ TRACEBACK COMPLETO:")
        print(traceback.format_exc())
        return jsonify({'error': f'Error interno del servidor: {str(e)}'}), 500

@admin_bp.route('/admin/parcelas/deshabilitar', methods=['POST'])
@jwt_required()
def deshabilitar_parcelas():
    """Deshabilitar parcelas seleccionadas - CORREGIDO"""
    try:
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return jsonify(permisos[0]), permisos[1]

        data = request.get_json()
        parcelas_ids = data.get('parcelas_ids', [])
        
        if not parcelas_ids:
            return jsonify({'error': 'No se proporcionaron parcelas para deshabilitar'}), 400
        
        # Verificar que las parcelas existan y no estén ocupadas - CONSULTA DIRECTA
        parcelas_ocupadas = []
        for parcela_id in parcelas_ids:
            parcela = Parcela.query.get(parcela_id)
            if not parcela:
                continue
                
            # Verificar si está ocupada - CONSULTA DIRECTA SIN RELACIONES
            solicitud_ocupada = db.session.query(SolicitudParcela).join(
                Solicitud, SolicitudParcela.solicitud_id == Solicitud.solicitud_id
            ).join(
                EstadoSolicitud, Solicitud.estado_solicitud_id == EstadoSolicitud.estado_solicitud_id
            ).filter(
                SolicitudParcela.parcela_id == parcela_id,
                EstadoSolicitud.nombre == 'Aprobada'
            ).first()
            
            if solicitud_ocupada:
                parcelas_ocupadas.append(parcela_id)
        
        if parcelas_ocupadas:
            return jsonify({
                'error': 'No se pueden deshabilitar parcelas ocupadas',
                'parcelas_ocupadas': parcelas_ocupadas
            }), 400
        
        # Deshabilitar las parcelas
        for parcela_id in parcelas_ids:
            parcela = Parcela.query.get(parcela_id)
            if parcela:
                parcela.habilitada = False
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(parcelas_ids)} parcelas deshabilitadas correctamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"❌ Error en deshabilitar_parcelas: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/admin/parcelas/habilitar', methods=['POST'])
@jwt_required()
def habilitar_parcelas():
    """Habilitar parcelas seleccionadas - CORREGIDO"""
    try:
        usuario = get_usuario_actual()
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return jsonify(permisos[0]), permisos[1]

        data = request.get_json()
        parcelas_ids = data.get('parcelas_ids', [])
        
        if not parcelas_ids:
            return jsonify({'error': 'No se proporcionaron parcelas para habilitar'}), 400
        
        # Habilitar las parcelas - MANERA DIRECTA
        for parcela_id in parcelas_ids:
            parcela = Parcela.query.get(parcela_id)
            if parcela:
                parcela.habilitada = True
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(parcelas_ids)} parcelas habilitadas correctamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"❌ Error en habilitar_parcelas: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/solicitudes/<int:solicitud_id>/modificar', methods=['PATCH'])
@jwt_required()
def modificar_informacion_puesto_route(solicitud_id):
    """RF13: Modificar información del puesto"""
    data = request.get_json()
    response, status = AdminController.modificar_informacion_puesto(solicitud_id, data)
    return jsonify(response), status

@admin_bp.route('/configuraciones/rubros', methods=['GET'])
@jwt_required()
def get_configuraciones_rubros_route():
    """RF17: Obtener configuraciones de rubros"""
    response, status = AdminController.get_configuraciones_rubros()
    return jsonify(response), status

@admin_bp.route('/configuraciones/rubros/<int:rubro_id>', methods=['PUT'])
@jwt_required()
def actualizar_configuracion_rubro_route(rubro_id):
    """RF17: Actualizar configuración de rubro"""
    data = request.get_json()
    response, status = AdminController.actualizar_configuracion_rubro(rubro_id, data)
    return jsonify(response), status

@admin_bp.route('/diversidad-rubros', methods=['GET'])
@jwt_required()
def get_diversidad_rubros_route():
    """RF14: Obtener diversidad por categorías"""
    response, status = AdminController.get_diversidad_rubros()
    return jsonify(response), status

# RUTAS EXISTENTES

@admin_bp.route('/solicitudes', methods=['GET'])
@jwt_required() 
def get_solicitudes_route():
    filtro = request.args.get('filtro_estado')
    busqueda = request.args.get('busqueda_termino')
    
    data, status = AdminController.get_solicitudes_dashboard(
        filtro_estado=filtro, 
        busqueda_termino=busqueda
    )
    return jsonify(data), status

@admin_bp.route('/solicitudes/<int:solicitud_id>/estado', methods=['PATCH'])
@jwt_required()
def actualizar_estado_route(solicitud_id):
    data = request.get_json()
    
    response, status = AdminController.actualizar_estado_solicitud(
        solicitud_id, 
        data
    )
    return jsonify(response), status

@admin_bp.route('/<int:solicitud_id>/fotos', methods=['GET'])
@jwt_required()
def obtener_fotos_solicitud(solicitud_id):
    try:
        usuario = get_usuario_actual()
        
        permisos = AdminController._check_admin_permissions(usuario)
        if not isinstance(permisos, Administrador):
            return jsonify(permisos[0]), permisos[1]
            
        solicitud = Solicitud.query.get(solicitud_id)
        if not solicitud:
            return jsonify({'msg': 'Solicitud no encontrada'}), 404
        
        fotos = SolicitudFoto.query.filter_by(solicitud_id=solicitud_id).all()
        
        fotos_respuesta = []
        for foto in fotos:
            foto_data = foto.to_dict()
            foto_data['image_url'] = foto.get_image_url()
            fotos_respuesta.append(foto_data)
        
        return jsonify({
            'fotos': fotos_respuesta
        }), 200
        
    except Exception as e:
        return jsonify({'msg': 'Error al obtener fotos', 'error': str(e)}), 500

@admin_bp.route('/estadisticas/usuarios', methods=['GET'])
@jwt_required()
def get_estadisticas_usuarios_route():
    try:
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')
        agrupacion = request.args.get('agrupacion', 'dia')

        if not fecha_inicio_str or not fecha_fin_str:
            return jsonify({'msg': 'Se requieren las fechas de inicio y fin'}), 400

        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()

        if agrupacion not in ['dia', 'semana', 'mes']:
            return jsonify({'msg': 'Agrupación no válida. Use: dia, semana o mes'}), 400

        data, status = AdminController.get_estadisticas_usuarios(fecha_inicio, fecha_fin, agrupacion)
        return jsonify(data), status

    except ValueError:
        return jsonify({'msg': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
    except Exception as e:
        return jsonify({'msg': 'Error interno del servidor', 'error': str(e)}), 500

@admin_bp.route('/estadisticas/usuarios/exportar-excel', methods=['GET'])
@jwt_required()
def exportar_estadisticas_excel_route():
    try:
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')
        agrupacion = request.args.get('agrupacion', 'dia')

        if not fecha_inicio_str or not fecha_fin_str:
            return jsonify({'msg': 'Se requieren las fechas de inicio y fin'}), 400

        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()

        if agrupacion not in ['dia', 'semana', 'mes']:
            return jsonify({'msg': 'Agrupación no válida. Use: dia, semana o mes'}), 400

        excel_buffer, status = AdminController.exportar_estadisticas_excel(fecha_inicio, fecha_fin, agrupacion)
        
        if status != 200:
            return excel_buffer, status

        nombre_archivo = f"reporte_usuarios_{fecha_inicio}_{fecha_fin}.xlsx"

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except ValueError:
        return jsonify({'msg': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
    except Exception as e:
        return jsonify({'msg': 'Error interno del servidor', 'error': str(e)}), 500

@admin_bp.route('/artesanos/exportar-pdf', methods=['GET'])
@jwt_required()
def exportar_artesanos_pdf_route():
    try:
        pdf_buffer, status = AdminController.exportar_artesanos_pdf()
        
        if status != 200:
            return pdf_buffer, status

        nombre_archivo = f"listado_artesanos_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/pdf'
        )

    except Exception as e:
        return jsonify({'msg': 'Error interno del servidor', 'error': str(e)}), 500

@admin_bp.route('/estadisticas/rubros', methods=['GET'])
@jwt_required()
def get_estadisticas_rubros_route():
    try:
        data, status = AdminController.get_estadisticas_rubros()
        return jsonify(data), status

    except Exception as e:
        return jsonify({'msg': 'Error interno del servidor', 'error': str(e)}), 500

@admin_bp.route('/estadisticas/rubros/todas', methods=['GET'])
@jwt_required()
def get_estadisticas_rubros_todas_route():
    try:
        data, status = AdminController.get_estadisticas_rubros_todas()
        return jsonify(data), status

    except Exception as e:
        return jsonify({'msg': 'Error interno del servidor', 'error': str(e)}), 500